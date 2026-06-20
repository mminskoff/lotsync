import uuid
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.adapters.inventory.base import InventoryAdapter
from app.schemas.normalized_vehicle import NormalizedVehicle

_STATUS_MAP = {
    "IN-STOCK": "available",
    "IN STOCK": "available",
    "AVAILABLE": "available",
    "SOLD": "sold",
}


def _cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _cell_price(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _normalize_status(raw: str | None) -> str:
    if not raw:
        return "available"
    key = raw.strip().upper()
    return _STATUS_MAP.get(key, raw.strip().lower())


def _row_to_vehicle(
    row: tuple, dealership_id: uuid.UUID, source_type: str
) -> NormalizedVehicle | None:
    vin = _cell_str(row[4])
    if not vin:
        return None

    price = _cell_price(row[24]) or _cell_price(row[12]) or _cell_price(row[13])
    photo_url = _cell_str(row[22])
    photos = [photo_url] if photo_url else []

    return NormalizedVehicle(
        vin=vin.upper(),
        stock_number=_cell_str(row[3]),
        year=_cell_int(row[5]),
        make=_cell_str(row[6]),
        model=_cell_str(row[7]),
        trim=_cell_str(row[9]),
        mileage=_cell_int(row[16]),
        price=price,
        price_type="internet_price",
        status=_normalize_status(_cell_str(row[29])),
        dealership_id=dealership_id,
        exterior_color=_cell_str(row[14]),
        interior_color=_cell_str(row[15]),
        body_style=_cell_str(row[26]),
        fuel_type=_cell_str(row[28]),
        transmission=_cell_str(row[10]),
        photos=photos,
        source_type=source_type,
    )


class NielsenDDCAdapter(InventoryAdapter):
    """Reads Nielsen DDC Excel workbooks (multi-tab dealer exports)."""

    def test_connection(self, config: dict) -> None:
        file_path = config.get("file_path")
        if not file_path:
            raise ValueError("config.file_path is required for Nielsen DDC imports")
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"Nielsen workbook not found: {path}")

        sheet_name = config.get("sheet_name")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    available = ", ".join(workbook.sheetnames)
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. Available sheets: {available}"
                    )
            elif len(workbook.sheetnames) == 0:
                raise ValueError("Workbook contains no sheets")
        finally:
            workbook.close()

    def fetch_inventory(
        self, dealership_id: uuid.UUID, config: dict
    ) -> list[NormalizedVehicle]:
        self.test_connection(config)
        path = Path(config["file_path"])
        sheet_name = config.get("sheet_name")
        source_type = config.get("source_type", "nielsen")

        workbook = load_workbook(path, read_only=True, data_only=True)
        vehicles: list[NormalizedVehicle] = []
        try:
            sheets = [sheet_name] if sheet_name else workbook.sheetnames
            for name in sheets:
                worksheet = workbook[name]
                rows = worksheet.iter_rows(min_row=2, values_only=True)
                for row in rows:
                    if not row or all(cell is None for cell in row):
                        continue
                    vehicle = _row_to_vehicle(row, dealership_id, source_type)
                    if vehicle:
                        vehicles.append(vehicle)
        finally:
            workbook.close()

        return vehicles
