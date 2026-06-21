"""Split Nielsen DDC workbook into one dealership (rooftop) per Excel tab.

Run from apps/api:
    PYTHONPATH=. .venv/bin/python scripts/setup_nielsen_rooftops.py

Requires the workbook at the path stored on the existing Nielsen inventory source,
or NIELSEN_WORKBOOK_PATH env / --workbook argument.
"""

from __future__ import annotations

import argparse
import re
import uuid
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.database import SessionLocal
from app.models.dealership import Dealership
from app.models.inventory_source import InventorySource
from app.models.organization import Organization
from app.models.vehicle import Vehicle
from app.services.inventory_sync_service import sync_inventory_source_now

NIELSEN_ORG_ID = uuid.UUID("00000000-0000-4000-8000-000000000001")
MORRISTOWN_DEALERSHIP_ID = uuid.UUID("bcc16cb3-dd58-44fa-8777-271735a4afb5")
DEFAULT_WORKBOOK = Path("/Users/mikey/Downloads/Nielsen DDC Files.xlsx")

SHEET_DEALERSHIP_OVERRIDES: dict[str, uuid.UUID] = {
    "Nielsen Ford of Morristown": MORRISTOWN_DEALERSHIP_ID,
}


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug or "rooftop"


def resolve_workbook_path(explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit)
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path

    db = SessionLocal()
    try:
        source = db.scalar(
            select(InventorySource).where(
                InventorySource.dealership_id == MORRISTOWN_DEALERSHIP_ID,
                InventorySource.source_type == "nielsen",
            )
        )
        if source and source.config_json.get("file_path"):
            path = Path(source.config_json["file_path"])
            if path.is_file():
                return path
    finally:
        db.close()

    if DEFAULT_WORKBOOK.is_file():
        return DEFAULT_WORKBOOK

    raise FileNotFoundError(
        "Nielsen workbook not found. Pass --workbook or set file_path on the inventory source."
    )


def list_workbook_sheets(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def ensure_organization(db) -> Organization:
    org = db.get(Organization, NIELSEN_ORG_ID)
    if org is None:
        org = Organization(
            id=NIELSEN_ORG_ID,
            name="Nielsen Auto Group",
            slug="nielsen-auto-group",
        )
        db.add(org)
        db.flush()
    return org


def upsert_rooftop_dealership(
    db, sheet_name: str, organization_id: uuid.UUID
) -> Dealership:
    slug = slugify(sheet_name)
    override_id = SHEET_DEALERSHIP_OVERRIDES.get(sheet_name)

    if override_id:
        dealership = db.get(Dealership, override_id)
        if dealership is None:
            raise RuntimeError(f"Expected dealership {override_id} for {sheet_name}")
        dealership.name = sheet_name
        dealership.slug = slug
        dealership.organization_id = organization_id
        dealership.status = "active"
        return dealership

    existing = db.scalar(select(Dealership).where(Dealership.slug == slug))
    if existing:
        existing.name = sheet_name
        existing.organization_id = organization_id
        existing.status = "active"
        return existing

    dealership = Dealership(
        id=uuid.uuid4(),
        name=sheet_name,
        slug=slug,
        organization_id=organization_id,
        status="active",
    )
    db.add(dealership)
    db.flush()
    return dealership


def upsert_inventory_source(
    db, dealership_id: uuid.UUID, sheet_name: str, workbook_path: Path
) -> InventorySource:
    config = {
        "file_path": str(workbook_path),
        "sheet_name": sheet_name,
        "mark_missing_off_lot": False,
    }
    source = db.scalar(
        select(InventorySource).where(
            InventorySource.dealership_id == dealership_id,
            InventorySource.source_type == "nielsen",
        )
    )
    if source:
        source.name = f"Nielsen DDC · {sheet_name}"
        source.config_json = config
        source.enabled = True
        return source

    source = InventorySource(
        id=uuid.uuid4(),
        dealership_id=dealership_id,
        source_type="nielsen",
        name=f"Nielsen DDC · {sheet_name}",
        config_json=config,
        enabled=True,
    )
    db.add(source)
    db.flush()
    return source


def clear_nielsen_inventory(db, dealership_ids: list[uuid.UUID]) -> int:
    if not dealership_ids:
        return 0
    result = db.execute(
        delete(Vehicle).where(
            Vehicle.dealership_id.in_(dealership_ids),
            Vehicle.source_type == "nielsen",
        )
    )
    return result.rowcount or 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Set up Nielsen multi-rooftop import")
    parser.add_argument("--workbook", help="Path to Nielsen DDC Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no DB writes")
    args = parser.parse_args()

    workbook_path = resolve_workbook_path(args.workbook)
    sheets = list_workbook_sheets(workbook_path)

    print(f"Workbook: {workbook_path}")
    print(f"Rooftops: {len(sheets)}")
    for sheet in sheets:
        print(f"  - {sheet}")

    if args.dry_run:
        return

    db = SessionLocal()
    try:
        ensure_organization(db)
        db.commit()

        dealerships: list[Dealership] = []
        for sheet in sheets:
            dealer = upsert_rooftop_dealership(db, sheet, NIELSEN_ORG_ID)
            dealerships.append(dealer)
        db.commit()

        dealer_ids = [dealer.id for dealer in dealerships]
        removed = clear_nielsen_inventory(db, dealer_ids)
        db.commit()
        print(f"\nCleared {removed} merged nielsen vehicles (test/manual rows kept)")

        totals: list[tuple[str, int]] = []
        for sheet, dealer in zip(sheets, dealerships, strict=True):
            source = upsert_inventory_source(db, dealer.id, sheet, workbook_path)
            db.commit()

            result = sync_inventory_source_now(db, dealer.id, source.id)
            totals.append((dealer.name, result.vehicles_imported))
            print(
                f"  {dealer.name}: {result.vehicles_imported} vehicles "
                f"({result.vehicles_created} new, {result.vehicles_updated} updated)"
            )

        print(f"\nDone — {len(dealerships)} rooftops under Nielsen Auto Group")
        print(f"Org id: {NIELSEN_ORG_ID}")
        print(f"Morristown id: {MORRISTOWN_DEALERSHIP_ID}")
        print(f"Total vehicles: {sum(count for _, count in totals)}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
